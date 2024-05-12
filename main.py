import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment

def filter_data(charge_number):
    dict[charge_number] = {}
    filtered_df = df[(df['Charge Number'] == charge_number)]
    for index, row in filtered_df.iterrows():
        if row['Name'] in dict[charge_number]:
            dict[charge_number][row['Name']] += int(row['Hours'])
        else:
            dict[charge_number][row['Name']] = int(row['Hours'])
    return dict

def write_comment(dictionary, charge_number, date):
    wb = load_workbook('Total_hours.xlsx')
    sheet = wb.active

    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value) == charge_number:
                r = cell.row
            if str(cell.value) == date:
                c = cell.column

    message = []
    for name, hours in dictionary[charge_number].items():
        message.append(f'{name} ({hours})')
    message = '\n'.join(message)
    sheet.cell(row = r, column = c).comment = Comment(message, 'Jose')
    wb.save('Total_hours.xlsx')

if __name__ == '__main__':
    print(f'Enter Charge Number: ')
    charge_number = input('>')
    print(f"Enter Date in the following format: 'YYYY-MM-DD 00:00:00'")
    date = input('>')
    
    charge_numbers = ['N123', 'N456', 'N789']
    df = pd.read_excel('Charging_.xlsx')
    dict = {}

    for cn in charge_numbers:
        dict = filter_data(charge_number=cn)
        # Date format: 'YYYY-MM-DD 00:00:00'
        write_comment(dictionary=dict, charge_number=cn, date=date)

    